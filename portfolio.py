import os
import win32com.client
import pandas as p
from pandas.tseries.offsets import *
import sys
from openpyxl import load_workbook
from openpyxl.utils import *
import pyexcel as ex
from xlsxwriter.utility import *
import xlwings as x
from collections import OrderedDict
from jpxlpy.convert import *

#today = p.datetime(2016, 10, 25)
today = p.datetime.today()
offset = BMonthEnd()
lme = offset.rollback(today)
dash_lme = lme.strftime('%d-%m-%Y')
us_lme = lme.strftime('%m-%d-%Y')

if sys.platform == 'darwin':
	InDir = '/Users/jamespardee/Google Drive/Dev/PortfolioCharacteristics/originals/'
	OutDir = '/Users/jamespardee/Google Drive/Dev/PortfolioCharacteristics/'
	InFile = dash_lme + ' PC.xls'
	file = us_lme + ' PC Aladdin.xlsm'
	ofile = os.path.normpath(OutDir + 'Portfolio_Stats ' + us_lme + '.xlsm')
	InPath = os.path.join(InDir, InFile)
	fpath = os.path.join(InDir, file)
	fpathlookthru = os.path.join(InDir + us_lme + ' LookThru.xlsm')
	portfolio = p.read_excel('/Users/jamespardee/Google Drive/Dev/PortfolioCharacteristics/originals/Ref.xlsx', sheetname='Portfolio', header=0, index_col=0)
	ticker = p.read_excel('/Users/jamespardee/Google Drive/Dev/PortfolioCharacteristics/originals/Ref.xlsx', sheetname='Ticker', header=0, index_col=0)
else:
	InDir = 'G:/COMMUNICATION/SUPPLEMENTAL REPORTING/PY/PortfolioCharacteristics/originals/'
	OutDir = 'G:/COMMUNICATION/SUPPLEMENTAL REPORTING/PY/PortfolioCharacteristics/'
	InFile = dash_lme + ' PC.xls'
	file = us_lme + ' PC Aladdin.xlsm'
	ofile = os.path.normpath(OutDir + 'Portfolio_Stats ' + us_lme + '.xlsm')
	InPath = os.path.join(InDir, InFile)
	fpath = os.path.join(InDir, file)
	fpathlookthru = os.path.join(InDir + us_lme + ' LookThru.xlsm')
	portfolio = p.read_excel('G:/COMMUNICATION/SUPPLEMENTAL REPORTING/PY/PortfolioCharacteristics/originals/Ref.xlsx', sheetname='Portfolio', header=0, index_col=0)
	ticker = p.read_excel('G:/COMMUNICATION/SUPPLEMENTAL REPORTING/PY/PortfolioCharacteristics/originals/Ref.xlsx', sheetname='Ticker', header=0, index_col=0)

ExcelConvert(InPath, ConvertTo='xlsm', RemoveOrg=False, OutFile=fpath)

def portfolio_list():
	port_list = []
	sheets = []
	book = load_workbook(os.path.join(fpath))
	for i in book.get_sheet_names():
		sheets.append(i)
		p = i[: i.find(' ')]
		port_list.append(p)
	return port_list, sheets

port_list, sheets = portfolio_list()

def to_dict(l):
	d = {}
	for i in l:
		d[i[0]] = i[1:]
	return d

widgets = ['Sector Profile', 'Moodys', 'S&P', 'Barclays', 'Stats', 'PortfolioYC', 'BenchmarkYC', 'ActiveYC', 'Issuer Weight (%)', 'Sector Weight (%)', 'Sector Duration', 'Issuer Duration', 'CDX']

def coords(sheet):
	dex = []
	adr = {}
	idx = {}
	book = load_workbook(os.path.join(fpath))
	ws = book.get_sheet_by_name(sheet)
	for col in ws.columns:
		for cell in col:
			dex.append(cell.value)
		break
	for i in widgets:
		idx[i] = [dex.index(i) + 5, 1]

	for k, v in idx.items():
		y = v[0]
		x = v[1]
		topL = [y, x]
		topl = xl_rowcol_to_cell(y - 1, x - 1)
		while ws.cell(row=y, column=x).value is not None and ws.cell(row=y, column=x).value is not '':
			x += 1
		#x = x - 1
		topR = [y, x - 1]
		topr = xl_rowcol_to_cell(y - 1, x - 1)

		x = v[1]
		y = v[0]
		while ws.cell(row=y, column=x).value is not None and ws.cell(row=y, column=x).value is not '':
			y += 1
		y = y - 1
		botL = [y, x]
		botl = xl_rowcol_to_cell(y - 1, x - 1)

		botR = [botL[0], topR[1]]
		botr = xl_rowcol_to_cell(botL[0] - 1, topR[1] - 1)
		idx[k] = [topL, topR, botL, botR]
		adr[k] = [topl, topr, botl, botr]
	return idx, adr

ex = x.App(visible=False)
wb = x.Book(fpath)

def adjustments(sheet, adr):
	sht = wb.sheets[sheet]
	cdx = sht.range(adr['CDX'][0] + ':' + adr['CDX'][3])
	cdx = cdx.resize(1, 5).value
	adj = {data: num if num != '' else 0.00 for data, num in zip(['Portfolio', 'OAS', 'SprdDuration', 'Notional%', 'CDX Yield'], cdx)}
	return adj

def KRD(sheet, adr):
	sht = wb.sheets[sheet]
	pyc = sht.range(adr['PortfolioYC'][0] + ':' + adr['PortfolioYC'][3]).options(ndim=2).value[0]
	byc = sht.range(adr['BenchmarkYC'][0] + ':' + adr['BenchmarkYC'][3]).options(ndim=2).value[0]
	ayc = sht.range(adr['ActiveYC'][0] + ':' + adr['ActiveYC'][3]).options(ndim=2).value[0]
	dat = OrderedDict([('Portfolio', pyc[1:]), ('Benchmark', byc[1:]), ('Active', ayc[1:])])
	inx = ['3M', '1Y', '2Y', '3Y', '5Y', '7Y', '10Y', '15Y', '20Y', '25Y', '30Y']
	krd = p.DataFrame(data=dat, index=inx)
	return krd

def chrtr(sheet, portfolio, adr):
	sht = wb.sheets[sheet]
	line = coordinate_to_tuple(adr['Stats'][0]), coordinate_to_tuple(adr['Stats'][3])
	pb = (line[0][0] - 1, line[0][1] + 1), (line[1][0] - 1, line[1][1])
	title = (line[0][0] - 2, line[0][1] + 1), (line[1][0] - 2, line[1][1])
	stat = (line[0][0], line[0][1] + 1), (line[1][0], line[1][1])

	title = sht.range(title[0], title[1]).options(ndim=2).value[0]
	pb = sht.range(pb[0], pb[1]).options(ndim=2).value[0]
	stat = sht.range(stat[0], stat[1]).options(ndim=2).value[0]

	pb[-1] = 'Portfolio'
	pb.append('Benchmark')
	stat.append("-")
	while title.count(None) >= 1:
		title.remove(None)
	title[title.index('')] = 'Portfolio Market Value ($)'
	port = []
	ben = []
	count = 1
	for s in stat:
		if count % 2 == 0:
			ben.append(s)
			count += 1
		else:
			port.append(s)
			count += 1
	dafm = p.DataFrame(data={'Portfolio': port, 'Benchmark': ben}, index=title, columns=['Portfolio', 'Benchmark'])
	LookThru = p.read_excel(fpathlookthru, sheetname='loktru', skiprows=6, skip_footer=3, parse_cols=4, index_col=0)
	dafm.loc['Coupon']['Portfolio'] = LookThru.loc[portfolio]['Coupon']
	return dafm

def ratings(sheet, adr):
	sht = wb.sheets[sheet]
	bar = sht.range(adr['Barclays'][0] + ':' + adr['Barclays'][3]).options(ndim=2).value[1:]
	mdy = sht.range(adr['Moodys'][0] + ':' + adr['Moodys'][3]).options(ndim=2).value[1:]
	sp = sht.range(adr['S&P'][0] + ':' + adr['S&P'][3]).options(ndim=2).value[1:]
	bardf = p.DataFrame(bar, columns=['idx', 'Portfolio', 'Benchmark'])
	spdf = p.DataFrame(sp, columns=['idx', 'Portfolio', 'Benchmark'])
	mdydf = p.DataFrame(mdy, columns=['idx', 'Portfolio', 'Benchmark'])
	bardf.replace(to_replace='', value=0, inplace=True)
	spdf.replace(to_replace='', value=0, inplace=True)
	mdydf.replace(to_replace='', value=0, inplace=True)
	for df in [bardf, spdf, mdydf]:
		df.loc[:, ['Portfolio', 'Benchmark']] = df.loc[:, ['Portfolio', 'Benchmark']] * 100
		i = list(range(len(df)))
		i.append(i.pop(0))
		df = df.reindex(i)
		yield df
	return bardf, spdf, mdydf

def CleanSector(df):
	sectors = ['Telecom', 'Banking', 'Midstream', 'Technology', 'Cable/Media', 'Automotive', 'Pharmaceuticals', 'REITs', 'Food and Beverage', 'Health Insurance', 'Metals and Mining', 'Airlines', 'Insurance', 'Home Construction', 'Building Materials', 'Environmental', 'Paper/Packaging', 'Transportation Services', 'Natural Gas', 'Lodging/Leisure', 'Industrial Other', 'Finance', 'Consumer Cyclical Services', 'Consumer Products', 'Tobacco', 'Aerospace and Defense', 'Chemicals', 'Railroads', 'Manufacturing/Machinery', 'Energy', 'Healthcare', 'Retailers', 'Utility', 'Non-Corporate']
	if len(df.columns) == 0:
		df['Portfolio'] = 0
		df['Benchmark'] = 0
		df['Active'] = 0
	else:
		pass
	try:
		df.columns = ['Portfolio', 'Benchmark', 'Active']
		df.rename_axis({'Non_Corprate': 'Non-Corporate', 'Finace': 'Finance'}, axis='index', inplace=True)
	finally:
		for i in sectors:
			if i in df.index:
				continue
			else:
				df.loc[i] = 0
	df.columns = ['Portfolio', 'Benchmark', 'Active']
	df.replace(to_replace='', value=0, inplace=True)
	df.sort_values(by='Active', axis=0, ascending=False, inplace=True)
	df.insert(0, 'Issuer', df.index)
	return df

def CleanIss(df):
	if len(df.columns) == 0:
		df['Portfolio'] = 0
		df['Benchmark'] = 0
		df['Active'] = 0
	else:
		pass
	df.columns = ['Portfolio', 'Benchmark', 'Active']
	for i in df.index:
		try:
			df.rename_axis({i: ticker.loc[i, 'Name']}, axis='index', inplace=True)
		except:
			df.rename_axis({i: ''}, axis='index', inplace=True)
	df.replace(to_replace='', value=0, inplace=True)
	df.sort_values(by='Active', axis=0, ascending=False, inplace=True)
	df.insert(0, 'Issuer', df.index)
	return df

def CredExp(sheet, adr):
	sht = wb.sheets[sheet]
	SectPer = sht.range(adr['Sector Weight (%)'][0] + ':' + adr['Sector Weight (%)'][3]).options(ndim=2).value[1:]
	SectDur = sht.range(adr['Sector Duration'][0] + ':' + adr['Sector Duration'][3]).options(ndim=2).value[1:]
	IssPer = sht.range(adr['Issuer Weight (%)'][0] + ':' + adr['Issuer Weight (%)'][3]).options(ndim=2).value[1:]
	IssDur = sht.range(adr['Issuer Duration'][0] + ':' + adr['Issuer Duration'][3]).options(ndim=2).value[1:]
	SectPer = to_dict(SectPer)
	SectDur = to_dict(SectDur)
	IssPer = to_dict(IssPer)
	IssDur = to_dict(IssDur)
	SectPerdf = p.DataFrame.from_dict(SectPer, orient='index')
	SectPerdf = SectPerdf * 100
	SectPerdf = CleanSector(SectPerdf)
	SectDurdf = p.DataFrame.from_dict(SectDur, orient='index')
	SectDurdf = CleanSector(SectDurdf)
	IssPerdf = p.DataFrame.from_dict(IssPer, orient='index')
	IssPerdf = IssPerdf * 100
	IssPerdf = CleanIss(IssPerdf)
	IssDurdf = p.DataFrame.from_dict(IssDur, orient='index')
	IssDurdf = CleanIss(IssDurdf)
	return SectPerdf, SectDurdf, IssPerdf, IssDurdf

def cpy():
	app = win32com.client.DispatchEx("Excel.Application")
	app.DisplayAlerts = False
	app.Visible = False
	temp = app.Workbooks.Open(InDir + 'temp.xlsm')
	for name in port_list:
		temp.Worksheets('temp').Copy(Before=temp.Worksheets('temp'))
		sht = temp.Worksheets('temp (2)')
		sht.Select()
		sht.Activate()
		sht.Name = name
	temp.Worksheets('temp').Delete()
	temp.SaveAs(ofile, FileFormat=52)
	temp.Close()
	app.Quit()
	del app

def xlsx():
	book = load_workbook(ofile, guess_types=False, data_only=True, keep_vba=True)
	writer = p.ExcelWriter(ofile, engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	for port, shet in zip(port_list, sheets):
		idx, adr = coords(shet)
		ws = book[port]
		krd = KRD(shet, adr)
		dafr = chrtr(shet, port, adr)
		bardf, spdf, mdydf = ratings(shet, adr)
		#SectPerdf, SectDurdf, IssPerdf, IssDurdf = CredExp(shet, adr)
		#SP1 = SectPerdf[:17]
		#SP2 = SectPerdf[17:]
		#SD1 = SectDurdf[:17]
		#SD2 = SectDurdf[17:]
		krd.to_excel(writer, ws.title, startcol=4, startrow=13, header=False, index=True)
		dafr.to_excel(writer, ws.title, startcol=1, startrow=1, header=False, index=False)
		bardf.to_excel(writer, ws.title, startcol=0, startrow=12, header=False, index=False)
		spdf.to_excel(writer, ws.title, startcol=0, startrow=25, header=False, index=False)
		mdydf.to_excel(writer, ws.title, startcol=0, startrow=37, header=False, index=False)
		SD1.to_excel(writer, ws.title, startcol=16, startrow=49, header=False, index=False)
		SD2.to_excel(writer, ws.title, startcol=21, startrow=49, header=False, index=False)
		IssPerdf.to_excel(writer, ws.title, startcol=26, startrow=49, header=False, index=False)
		IssDurdf.to_excel(writer, ws.title, startcol=31, startrow=49, header=False, index=False)
		for c in [(1, 2), (12, 2), (25, 2), (37, 2), (49, 18), (49, 23), (49, 28), (49, 33)]:
			ws.cell(row=c[0], column=c[1], value=portfolio.loc[port, 'Name'])
		for c in [(1, 3), (12, 3), (25, 3), (37, 3), (49, 19), (49, 24), (49, 29), (49, 34)]:
			ws.cell(row=c[0], column=c[1], value=portfolio.loc[port, 'Benchmark'])
		ws.cell(coordinate='C22', value='=SUM(C13:C21)')
		ws.cell(coordinate='B22', value='=SUM(B13:B21)')
		ws.cell(coordinate='B35', value='=SUM(B26:B34)')
		ws.cell(coordinate='C35', value='=SUM(C26:C34)')
		ws.cell(coordinate='B47', value='=SUM(B38:B46)')
		ws.cell(coordinate='C47', value='=SUM(C38:C46)')
		ws.cell(coordinate='W67', value='=SUM(W50:W66,R50:R66)')
		ws.cell(coordinate='X67', value='=SUM(X50:X66,S50:S66)')
		ws.cell(coordinate='Y67', value='=SUM(Y50:Y66,T50:T66)')
	book.save(ofile)

def cp_chart(path_to=ofile, path_from=InDir + 'temp.xlsm', ws_from='temp'):
	app = win32com.client.DispatchEx("Excel.Application")
	app.Visible = False
	app.DisplayAlerts = False
	wbfrm = app.Workbooks.Open(path_from)
	wbfrm.Worksheets(ws_from).ChartObjects(1).Activate()
	wbfrm.Worksheets(ws_from).ChartObjects(1).Select()
	wbfrm.Worksheets(ws_from).ChartObjects(1).Copy()
	wbto = app.Workbooks.Open(path_to)
	for sht in wbto.Worksheets:
		sht.Activate()
		sht.Select()
		sht.Range("E2").Select()
		sht.Paste()
	for i in wbto.Worksheets:
		i.Activate()
		i.Select()
		for cs in i.ChartObjects():
			cs.Activate()
			cs.Select()
			cs.Chart.SeriesCollection(1).Select()
			cs.Chart.SetSourceData(Source=i.Range("=$E$13:$H$24"))
	wbfrm.Close()
	wbto.SaveAs(ofile, FileFormat=52)
	wbto.Close()
	app.Quit()
	del app

def cp_porfile(path_to=ofile, path_from=fpath):
	app = win32com.client.DispatchEx("Excel.Application")
	#app.Visible = False
	#app.DisplayAlerts = False
	wbfrm = app.Workbooks.Open(path_from, ReadOnly=1)
	wbto = app.Workbooks.Open(path_to)
	for sht, prt in zip(sheets, port_list):
		idx, adr = coords(sht)
		s = wbfrm.Worksheets(sht)
		s.Activate()
		try:
			app.Application.Run('C:\\Users\\jpardee\\AppData\\Roaming\\Microsoft\\Excel\\XLSTART\\PERSONAL.XLSB!Module1.FindSectorProfile')
		except:
			app.Run(Macro='PERSONAL.XLSB!Module1.FindSectorProfile')
		s.Range(xl_rowcol_to_cell(idx['Sector Profile'][0][0] - 3, 0) + ':' + xl_rowcol_to_cell(idx['Sector Profile'][3][0] - 1, idx['Sector Profile'][3][1] - 1)).Select()
		s.Range(xl_rowcol_to_cell(idx['Sector Profile'][0][0] - 3, 0) + ':' + xl_rowcol_to_cell(idx['Sector Profile'][3][0] - 1, idx['Sector Profile'][3][1] - 1)).Copy()
		sto = wbto.Worksheets(prt)
		sto.Activate()
		sto.Range("A50").Select()
		sto.Paste()
		sto.Columns('A:A').AutoFit()
		sto.Rows('49:49').AutoFit()
		sto.Range('A1').Select()
	wbto.Worksheets(1).Select()
	wbto.Worksheets(1).Activate()
	wbto.SaveAs(path_to, FileFormat=52)
	wbto.Close()
	wbfrm.Close()
	app.Quit()
	del app


#port_list, sheets = portfolio_list()

#idx, adr = coords(sheets[0])
#krd = KRD(sheets[0], idx)
#dafm = chrtr(sheets[0], port_list[0], idx)
#bardf, spdf, mdydf = ratings(sheets[0], idx)
#SectPerdf, SectDurdf, IssPerdf, IssDurdf = CredExp(sheets[0], idx)

cpy()
xlsx()
cp_chart(path_to=ofile, path_from=InDir + 'temp.xlsm', ws_from='temp')
cp_porfile()

wb.save()
